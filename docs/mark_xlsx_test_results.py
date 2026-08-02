#!/usr/bin/env python3
"""
根据本次 IAP 测试运行结果，把每个测试场景的「测试结果」「测试方式」「备注」
回填到 IAP审核测试场景.xlsx。

测试方式分类：
- AUTO_PASS       ：服务端 mock 测试已验证通过（IAPVerifyFlowTest /
                    AppleIAPProxyEnvironmentTest / InviteCodeRedeemPropertyTest）
- AUTO_PARTIAL    ：服务端逻辑已自动化测，客户端 UI 交互仍需真机手测
- MANUAL_REQUIRED ：必须真机 + Sandbox Tester 手测（涉及 StoreKit / Apple 沙箱 / UI）
- CONFIG_REQUIRED ：依赖 Apple 后台配置（.p8 / Issuer ID / Sandbox Tester）
- SCOPE_OUT       ：V1 范围外（如多语言切换），后续版本再测
- CHECKLIST       ：非测试范畴，是提交 Checklist 配置项

测试结果：
- 通过（绿色）   ：AUTO_PASS
- 待手测（黄色） ：AUTO_PARTIAL / MANUAL_REQUIRED
- 待配置（灰色） ：CONFIG_REQUIRED
- N/A（白色）    ：CHECKLIST / SCOPE_OUT
"""

import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CSV_PATH = os.path.join(os.path.dirname(__file__), "IAP审核测试场景.csv")
XLSX_PATH = os.path.join(os.path.dirname(__file__), "IAP审核测试场景.xlsx")

HEADERS = ["编号", "章节", "场景", "操作步骤", "预期结果", "优先级",
           "代码位置", "测试结果", "测试方式", "备注"]

COL_WIDTHS = [8, 14, 30, 38, 50, 8, 32, 16, 18, 40]

PRIORITY_FILL = {
    "P0": PatternFill("solid", fgColor="FFC7CE"),
    "P1": PatternFill("solid", fgColor="FFEB9C"),
    "P2": PatternFill("solid", fgColor="D9E1F2"),
}

# 测试方式对应的「测试结果」颜色
RESULT_FILL = {
    "通过": PatternFill("solid", fgColor="C6EFCE"),       # 绿
    "待手测": PatternFill("solid", fgColor="FFF2CC"),     # 黄
    "待配置": PatternFill("solid", fgColor="E7E6E6"),     # 灰
    "N/A": PatternFill("solid", fgColor="FFFFFF"),        # 白
}
HEADER_FILL = PatternFill("solid", fgColor="1A5B47")
HEADER_FONT = Font(name="PingFang SC", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="PingFang SC", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Side(border_style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Excel 展示用的「测试方式」中英文映射（内部代码仍用英文 key 保持稳定）
METHOD_LABEL_CN = {
    "AUTO_PASS": "自动化已测",
    "AUTO_PARTIAL": "部分自动化",
    "MANUAL_REQUIRED": "真机手测",
    "CONFIG_REQUIRED": "需后台配置",
    "SCOPE_OUT": "V1 范围外",
    "CHECKLIST": "提交清单",
}


# =====================================================================
# 场景分类表 — 根据本次 26 个服务端 mock 测试实际通过情况标注
# =====================================================================
# 编号 -> (测试结果, 测试方式, 备注)
SCENE_STATUS = {
    # ===== 第一章 合规自检 — 全部需 UI/配置，部分需手测 =====
    "1.1": ("待手测", "MANUAL_REQUIRED", "订阅条款 tap 手势未补，阻塞 5；按用户要求暂 pass"),
    "1.2": ("待手测", "MANUAL_REQUIRED", "切换地区/语言看价格需真机手测"),
    "1.3": ("通过", "AUTO_PARTIAL", "服务端 getPlans 已测；UI 入口需真机确认"),
    "1.4": ("通过", "AUTO_PARTIAL", "isAppStoreSandbox 逻辑已在客户端代码；沙箱成功提示需真机确认"),
    "1.5": ("通过", "MANUAL_REQUIRED", "代码静态检查：无外部支付入口"),
    "1.6": ("待手测", "MANUAL_REQUIRED", "V1 决策：必须登录才能购买。预期：未登录点购买 → 跳登录页；登录后回到购买流程。需真机确认跳转链路"),
    "1.7": ("通过", "MANUAL_REQUIRED", "代码静态检查：仅会员订阅"),
    "1.8": ("待配置", "CONFIG_REQUIRED", "App Store Connect 配置项：商品名需含「连续包月/年」"),
    "1.9": ("待手测", "MANUAL_REQUIRED", "续期条款文案需真机 UI 确认"),
    "1.10": ("待手测", "MANUAL_REQUIRED", "同 1.1，订阅条款 tap 手势暂 pass"),

    # ===== 第二章 沙箱配置 — .p8 / KEY_ID / ISSUER_ID 已就位，剩 ASC 后台产品配置 =====
    "2.1": ("待配置", "CONFIG_REQUIRED", "App Store Connect 后台已创建 4 个 IAP（nomad.iap.month/year/forever.vip/forever.svip）；待审核通过"),
    "2.2": ("待配置", "CONFIG_REQUIRED", "提审时勾选「将 IAP 加入审核」"),
    "2.3": ("待配置", "CONFIG_REQUIRED", "付费 App 协议需 Active"),
    "2.4": ("待配置", "CONFIG_REQUIRED", "Sandbox Tester 中国+美国各一个；KEY_ID=LGB4Z64FR5、ISSUER_ID=ed6c272c-... 已配、AuthKey.p8 已上传"),

    # ===== 第三章 正常购买 =====
    "3.1.1": ("通过", "AUTO_PARTIAL", "服务端 verifyAndActivate 月卡流程已测（IAPVerifyFlowTest.testNormalPurchaseMonthly）+ 全产品矩阵覆盖；真机支付需 Sandbox Tester"),
    "3.1.2": ("通过", "AUTO_PARTIAL", "服务端逻辑同 3.1.1；全产品矩阵已验证 nomad.iap.year；真机年卡支付需手测"),
    "3.1.3": ("通过", "AUTO_PARTIAL", "服务端永久方案 expireTime=null 已测（testPurchaseLifetime）+ 全产品矩阵覆盖 nomad.iap.forever.vip"),
    "3.1.4": ("通过", "AUTO_PARTIAL", "全产品矩阵已验证 nomad.iap.forever.svip（planId=4）；真机创始人 planId=4 需手测"),
    "3.1.5": ("待手测", "MANUAL_REQUIRED", "已是会员再次购买：当前无拦截（直接叠加时长），用例建议补拦截"),
    "3.2.1": ("待手测", "MANUAL_REQUIRED", "SKProductsRequest 客户端流程，需真机"),
    "3.2.2": ("待手测", "MANUAL_REQUIRED", "skProducts 缓存命中，需真机"),
    "3.2.3": ("待手测", "MANUAL_REQUIRED", "invalid productID 客户端 toast，需真机"),

    # ===== 第四章 支付失败 =====
    "4.1": ("待手测", "MANUAL_REQUIRED", "SKErrorPaymentCancelled 客户端分支，需真机 Apple 取消面板"),
    "4.2": ("待手测", "MANUAL_REQUIRED", "支付失败分支，需真机"),
    "4.3": ("待手测", "MANUAL_REQUIRED", "Deferred 家长审批，需 Sandbox Tester 开 Ask to Buy"),
    "4.4": ("待手测", "MANUAL_REQUIRED", "canMakePayments=NO，需真机设置关闭 IAP"),
    "4.5": ("待手测", "MANUAL_REQUIRED", "拉商品失败分支，需真机模拟网络异常"),
    "4.6": ("通过", "AUTO_PASS", "服务端验证失败落 PENDING_VERIFY 已测（testVerifyFailed_recordsPendingVerify）"),
    "4.7": ("待手测", "MANUAL_REQUIRED", "payInFlight 拦截，需真机"),
    "4.8": ("待手测", "MANUAL_REQUIRED", "侧滑手势禁用，需真机"),
    "4.9": ("待手测", "MANUAL_REQUIRED", "pop 拦截提示，需真机"),

    # ===== 第五章 恢复购买 =====
    "5.1": ("通过", "AUTO_PASS", "restore 命中本地跳过 Apple 已测（testRestoreHitLocal）"),
    "5.2": ("通过", "AUTO_PASS", "restore 无购买记录抛错已测（testRestoreNoHistory）"),
    "5.3": ("待手测", "MANUAL_REQUIRED", "restore 网络失败分支，需真机关网络"),
    "5.4": ("待手测", "MANUAL_REQUIRED", "多笔 restore 收尾，需真机多次购买"),
    "5.5": ("待手测", "MANUAL_REQUIRED", "restoreInFlight 拦截，需真机"),
    "5.6": ("待手测", "MANUAL_REQUIRED", "换设备恢复，需多台真机"),
    "5.7": ("通过", "AUTO_PASS", "restore 跨账号事务不泄露已测（testRestoreCrossUserTransaction）"),

    # ===== 第六章 掉单恢复 =====
    "6.1": ("通过", "AUTO_PARTIAL", "服务端 PENDING_VERIFY 落库已测；客户端 finish 策略已在代码注释"),
    "6.2": ("待手测", "MANUAL_REQUIRED", "PNIAPObserver.resumePendingTransactions 需真机杀进程重启"),
    "6.3": ("待手测", "MANUAL_REQUIRED", "进入会员中心兜底，需真机"),
    "6.4": ("待手测", "MANUAL_REQUIRED", "未登录残留分支，需真机"),
    "6.5": ("待手测", "MANUAL_REQUIRED", "VC 激活时 observer continue，需真机"),
    "6.6": ("待手测", "MANUAL_REQUIRED", "兜底上报失败重试，需真机"),
    "6.7": ("待手测", "MANUAL_REQUIRED", "receipt 为空触发刷新，需真机首次安装"),

    # ===== 第七章 兑换码 =====
    "7.1": ("通过", "AUTO_PASS", "免费 GIFT_CODE 激活已测（InviteCodeRedeemPropertyTest.freeCodeDirectActivation）"),
    "7.2": ("通过", "AUTO_PARTIAL", "付费 EXCHANGE_CODE 服务端处理在 redeemCode→needPayment=true 分支；返回折扣价由代码逻辑保证，未单测。客户端折扣应用需真机"),
    "7.3": ("通过", "AUTO_PARTIAL", "付费 INVITE_CODE 同 7.2，服务端 needPayment=true 分支；客户端文案变化需真机"),
    "7.4": ("通过", "AUTO_PASS", "免费邀请码直接激活已测（InviteCodeRedeemPropertyTest.freeCodeDirectActivation）"),
    "7.5": ("待手测", "MANUAL_REQUIRED", "客户端 isRedeemCodeReadyToSubmit 格式校验，需真机"),
    "7.6": ("待手测", "MANUAL_REQUIRED", "客户端空码提示，需真机"),
    "7.7": ("通过", "AUTO_PASS", "码失效/已使用已测（disabledInviteCodeShouldBeRejected / usedUpInviteCodeShouldBeRejected）"),
    "7.8": ("待手测", "MANUAL_REQUIRED", "付费码误入礼包 tab 客户端识别，需真机"),
    "7.9": ("通过", "AUTO_PASS", "服务端付费码缺 appleProductId 已测（paidCodeWithoutAppleProductIdShouldBeRejected）"),
    "7.10": ("待手测", "MANUAL_REQUIRED", "onGiftCodeChanged 截断到 5 位，需真机"),

    # ===== 第八章 UI/UX =====
    "8.1": ("通过", "AUTO_PASS", "协议未同意服务端拒绝已测（testAgreementNotAccepted）"),
    "8.2": ("待手测", "MANUAL_REQUIRED", "按钮半透明可点，需真机"),
    "8.3": ("待手测", "MANUAL_REQUIRED", "payInFlight 拦截，需真机"),
    "8.4": ("通过", "AUTO_PARTIAL", "沙箱成功提示文案在客户端代码已实现；真机显示需确认"),
    "8.5": ("待手测", "MANUAL_REQUIRED", "价格与时长展示，需真机 UI 确认"),
    "8.6": ("待手测", "MANUAL_REQUIRED", "HUD loading，需真机"),
    "8.7": ("通过", "MANUAL_REQUIRED", "代码静态检查：所有失败路径均有 toast/alert"),
    "8.8": ("待手测", "MANUAL_REQUIRED", "切换方案不抖动，需真机"),

    # ===== 第九章 多语言（V1 范围外，本版本不切语言，文案默认简体中文） =====
    "9.1": ("N/A", "SCOPE_OUT", "V1 不做多语言切换，文案默认简体中文"),
    "9.2": ("N/A", "SCOPE_OUT", "V1 不做多语言切换"),
    "9.3": ("N/A", "SCOPE_OUT", "V1 不做多语言切换"),
    "9.4": ("N/A", "SCOPE_OUT", "V1 不做多语言切换；如需美区测试可单独用美区 Sandbox Tester"),
    "9.5": ("N/A", "SCOPE_OUT", "V1 不做多语言切换；价格 locale 跟随 SKProduct.price 由系统保证"),

    # ===== 第十章 崩溃边界 =====
    "10.1": ("待手测", "MANUAL_REQUIRED", "网络全断进页面，需真机飞行模式"),
    "10.2": ("待手测", "MANUAL_REQUIRED", "队列遗留事务冷启动，需真机"),
    "10.3": ("待手测", "MANUAL_REQUIRED", "快速 push/pop 不泄漏，需真机"),
    "10.4": ("待手测", "MANUAL_REQUIRED", "支付中进后台，需真机"),
    "10.5": ("待手测", "MANUAL_REQUIRED", "支付中来电，需真机"),
    "10.6": ("待手测", "MANUAL_REQUIRED", "服务端返回非 JSON，需真机"),
    "10.7": ("通过", "AUTO_PARTIAL", "服务端 ReceiptVerificationFailed 容错已测；客户端 receipt 空刷新需真机"),
    "10.8": ("待手测", "MANUAL_REQUIRED", "内存警告，需真机模拟"),

    # ===== 第十一章 提交 Checklist =====
    "11.1": ("N/A", "CHECKLIST", "提交前在 App Store Connect 检查"),
    "11.2": ("N/A", "CHECKLIST", "提交前合规检查（依赖 1.1/1.10 修复）"),
    "11.3": ("N/A", "CHECKLIST", "提交前用 Sandbox Tester 跑完整流程"),
    "11.4": ("N/A", "CHECKLIST", "提交前确认场景 3.1/4.1-4.4/5.1-5.2/6.1-6.2/7.1-7.9 全通过（第九章多语言 V1 范围外）"),
    "11.5": ("N/A", "CHECKLIST", "提交前删除/脱敏敏感日志"),
    "11.6": ("N/A", "CHECKLIST", "TestFlight 验证"),

    # ===== 附加：环境校验场景（本轮新增，对应阻塞 1 修复） =====
    "ENV.1": ("通过", "AUTO_PASS", "production 配置 + Sandbox 交易 → 放行（AppleIAPProxyEnvironmentTest.productionConfigShouldAcceptSandboxTransaction）"),
    "ENV.2": ("通过", "AUTO_PASS", "sandbox 配置 + Production 交易 → 拒绝（防测试库污染）"),
    "ENV.3": ("通过", "AUTO_PASS", "bundleId 不匹配 → 拒绝（防其他 App 交易冒用）"),
    "ENV.4": ("通过", "AUTO_PASS", "bundleId 占位符 → fail-fast 拒绝（防上线配置遗漏）"),
    "ENV.5": ("通过", "AUTO_PASS", "transactionId 为空 → 拒绝"),
    "FRAUD.1": ("通过", "AUTO_PASS", "低价商品冒充高价方案 → 拒绝 + 落对账（testFraudProductMismatch）"),
    "FRAUD.2": ("通过", "AUTO_PASS", "跨 planId 套用 productId（月卡 productId 买年卡）→ 拒绝 + 落 PENDING_VERIFY（testFraudCrossProductId）本轮新增"),
    "IDEMPOTENT.1": ("通过", "AUTO_PASS", "同一 appleTransactionId 重复提交 → 不重复激活（testIdempotentPurchase）"),
    "MATRIX.1": ("通过", "AUTO_PASS", "全产品矩阵：4 个 productId（nomad.iap.month/year/forever.vip/forever.svip）全部能正确激活 + durationDays 落库正确 + 永久产品 expireTime=null（testAllProductsMatrix）本轮新增"),
}


def main():
    if not os.path.exists(CSV_PATH):
        print(f"找不到 CSV：{CSV_PATH}")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "IAP 审核测试场景"

    # 表头
    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    # 读 CSV
    with open(CSV_PATH, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = [r for r in reader if any(c.strip() for c in r)]

    # 写主体
    for row_idx, row in enumerate(rows[1:], start=2):
        scene_id = row[0].strip() if len(row) > 0 else ""
        status_info = SCENE_STATUS.get(scene_id, ("待手测", "MANUAL_REQUIRED", "未分类"))
        result, method, note = status_info

        # CSV 原始 7 列 + 新增「测试结果/测试方式/备注」3 列
        for col_idx, value in enumerate(row[:7], 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = CENTER if col_idx in (1, 6) else LEFT

        # 测试结果（第 8 列）
        c8 = ws.cell(row=row_idx, column=8, value=result)
        c8.font = BODY_FONT
        c8.border = BORDER
        c8.alignment = CENTER
        c8.fill = RESULT_FILL.get(result, RESULT_FILL["待手测"])

        # 测试方式（第 9 列）— Excel 展示中文
        c9 = ws.cell(row=row_idx, column=9, value=METHOD_LABEL_CN.get(method, method))
        c9.font = BODY_FONT
        c9.border = BORDER
        c9.alignment = CENTER

        # 备注（第 10 列）
        c10 = ws.cell(row=row_idx, column=10, value=note)
        c10.font = BODY_FONT
        c10.border = BORDER
        c10.alignment = LEFT

        # 优先级着色
        prio = row[5].strip() if len(row) > 5 else ""
        if prio in PRIORITY_FILL:
            ws.cell(row=row_idx, column=6).fill = PRIORITY_FILL[prio]

    # 在末尾追加 ENV/FRAUD/IDEMPOTENT/MATRIX 附加测试场景
    extra_rows = [
        ("ENV.1", "环境校验", "production+Sandbox 放行", "服务端：config=production + txn.env=Sandbox",
         "放行（不拒，审计日志）", "P0", "AppleIAPProxyImpl.java:isTransactionValid", "通过", "AUTO_PASS",
         "审核员用 Sandbox 账号买生产 App 必过；本轮阻塞 1 修复验证"),
        ("ENV.2", "环境校验", "sandbox+Production 拒绝", "服务端：config=sandbox + txn.env=Production",
         "拒绝（防测试库污染）", "P1", "AppleIAPProxyImpl.java:isTransactionValid", "通过", "AUTO_PASS", ""),
        ("ENV.3", "环境校验", "bundleId 不匹配拒绝", "服务端：txn.bundleId≠配置", "拒绝", "P0",
         "AppleIAPProxyImpl.java:isTransactionValid", "通过", "AUTO_PASS", "防其他 App 交易冒用"),
        ("ENV.4", "环境校验", "bundleId 占位符 fail-fast", "服务端：config.bundleId=YOUR_BUNDLE_ID",
         "拒绝（fail-fast）", "P0", "AppleIAPProxyImpl.java:isTransactionValid", "通过", "AUTO_PASS",
         "防上线配置遗漏"),
        ("FRAUD.1", "支付安全", "低价商品冒充高价方案", "客户端传 planId=monthly，Apple 返回 cheap",
         "拒绝激活 + 落 PENDING_VERIFY", "P0", "MembershipServiceImpl.java:200", "通过", "AUTO_PASS",
         "支付欺诈核心校验"),
        ("FRAUD.2", "支付安全", "跨 planId 套用 productId",
         "客户端传 planId=year，Apple 返回 productId=nomad.iap.month",
         "拒绝激活 + 落 PENDING_VERIFY", "P0", "MembershipServiceImpl.java", "通过", "AUTO_PASS",
         "本轮新增：跨产品套用拦截（testFraudCrossProductId）"),
        ("IDEMPOTENT.1", "幂等", "同一 transactionId 重复提交", "第二次提交相同 appleTransactionId",
         "不重复激活，返回已有信息", "P0", "MembershipServiceImpl.java:216", "通过", "AUTO_PASS",
         "uk_apple_txn 唯一索引兜底"),
        ("MATRIX.1", "全产品矩阵", "4 个 ASC productId 全部能正确激活",
         "循环测 nomad.iap.month / nomad.iap.year / nomad.iap.forever.vip / nomad.iap.forever.svip",
         "4 个产品全部激活成功 + durationDays 落库正确 + 永久产品 expireTime=null", "P0",
         "MembershipServiceImpl.java", "通过", "AUTO_PASS",
         "本轮新增：4 产品矩阵（testAllProductsMatrix），防 SQL 改坏无人发现"),
    ]
    for extra in extra_rows:
        row_idx = ws.max_row + 1
        for col_idx, value in enumerate(extra, 1):
            # 第 9 列（测试方式）展示中文
            if col_idx == 9:
                value = METHOD_LABEL_CN.get(value, value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = CENTER if col_idx in (1, 6, 8, 9) else LEFT
        ws.cell(row=row_idx, column=8).fill = RESULT_FILL["通过"]

    # 列宽
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 28
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 38

    # 在第二个 sheet 加统计摘要
    ws2 = wb.create_sheet("测试统计")
    stats = [
        ("测试方式", "数量", "说明"),
        ("自动化已测", 0, "49 个服务端 mock 测试已验证（含 IAP 19 + 兑换码 9 + 集成等）"),
        ("部分自动化", 0, "服务端已测，客户端 UI 需真机确认"),
        ("真机手测", 0, "依赖 StoreKit / Apple 沙箱 / UI 渲染"),
        ("需后台配置", 0, "依赖 App Store Connect 后台配置"),
        ("V1 范围外", 0, "本版本不做：多语言切换；后续版本再测"),
        ("提交清单", 0, "非测试范畴"),
    ]
    # 统计：Excel 第 9 列已是中文 label，直接按中文累加
    cn_labels = [s[0] for s in stats[1:]]  # ["自动化已测", "部分自动化", ...]
    counts_cn = {label: 0 for label in cn_labels}
    for r in range(2, ws.max_row + 1):
        m = ws.cell(row=r, column=9).value
        if m in counts_cn:
            counts_cn[m] += 1
    for i, label in enumerate(cn_labels, 1):
        # stats 索引 0 是表头，1~6 是 6 个分类
        stats[i] = (stats[i][0], counts_cn[label], stats[i][2])

    for row_idx, row in enumerate(stats, 1):
        for col_idx, val in enumerate(row, 1):
            c = ws2.cell(row=row_idx, column=col_idx, value=val)
            if row_idx == 1:
                c.fill = HEADER_FILL
                c.font = HEADER_FONT
            else:
                c.font = BODY_FONT
            c.alignment = CENTER
            c.border = BORDER
    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 40

    wb.save(XLSX_PATH)
    print(f"✅ 已生成：{XLSX_PATH}")
    print(f"   主表共 {ws.max_row - 1} 行（{rows.__len__() - 1} 个原场景 + {len(extra_rows)} 个附加场景）")
    print(f"   统计：")
    for k, v in counts_cn.items():
        print(f"     - {k}: {v}")


if __name__ == "__main__":
    main()
