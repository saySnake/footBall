#!/usr/bin/env python3
"""
将 IAP 审核测试场景 CSV 转为格式化的 Excel（.xlsx）。

特性：
- 表头加粗 + 深色背景 + 白字
- 冻结首行
- 自动列宽（按内容长度）
- 「优先级」列按 P0/P1/P2 着色
- 「测试结果」列高亮可填

依赖：pip3 install openpyxl
"""

import csv
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("缺少依赖：pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

CSV_PATH = os.path.join(os.path.dirname(__file__), "IAP审核测试场景.csv")
XLSX_PATH = os.path.join(os.path.dirname(__file__), "IAP审核测试场景.xlsx")

HEADERS = ["编号", "章节", "场景", "操作步骤", "预期结果", "优先级", "代码位置", "测试结果", "备注"]

# 列宽（字符数）
COL_WIDTHS = [8, 14, 30, 38, 50, 8, 32, 14, 30]

PRIORITY_FILL = {
    "P0": PatternFill("solid", fgColor="FFC7CE"),  # 浅红，必须通过
    "P1": PatternFill("solid", fgColor="FFEB9C"),  # 浅黄，重要
    "P2": PatternFill("solid", fgColor="D9E1F2"),  # 浅蓝，普通
}

HEADER_FILL = PatternFill("solid", fgColor="1A5B47")  # 项目主色
HEADER_FONT = Font(name="PingFang SC", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="PingFang SC", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Side(border_style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def main():
    if not os.path.exists(CSV_PATH):
        print(f"找不到 CSV：{CSV_PATH}", file=sys.stderr)
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "IAP 审核测试场景"

    # 写表头
    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    # 读 CSV 并写入
    with open(CSV_PATH, "r", encoding="utf-8") as f:
        # 跳过空行
        reader = csv.reader(f)
        rows = [r for r in reader if any(c.strip() for c in r)]

    # 第一行是 CSV 表头，跳过
    for row_idx, row in enumerate(rows[1:], start=2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.border = BORDER
            # 编号/优先级/测试结果居中，其它左对齐
            cell.alignment = CENTER if col_idx in (1, 6, 8) else LEFT
        # 优先级着色
        prio = row[5].strip() if len(row) > 5 else ""
        if prio in PRIORITY_FILL:
            ws.cell(row=row_idx, column=6).fill = PRIORITY_FILL[prio]
        # 测试结果列底色（提示填写）
        ws.cell(row=row_idx, column=8).fill = PatternFill("solid", fgColor="FFF2CC")

    # 列宽
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动筛选
    ws.auto_filter.ref = ws.dimensions

    # 行高
    ws.row_dimensions[1].height = 28
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 32

    wb.save(XLSX_PATH)
    print(f"✅ 已生成：{XLSX_PATH}")
    print(f"   共 {ws.max_row - 1} 个测试场景")


if __name__ == "__main__":
    main()
